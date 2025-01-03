import pandas as pd
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill

def choose_file():
    """让用户选择文件"""
    root = tk.Tk()
    root.withdraw()  # 隐藏主窗口

    file_path = filedialog.askopenfilename(
        title="选择数据文件",
        filetypes=[("Excel 文件", "*.xlsx *.xls"), ("CSV 文件", "*.csv"), ("所有文件", "*.*")]
    )
    return file_path

def load_data(file_path):
    """加载数据"""
    try:
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        elif file_path.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file_path)
        else:
            messagebox.showerror("错误", "不支持的文件类型")
            return None
    except Exception as e:
        messagebox.showerror("错误", f"加载数据失败: {e}")
        return None
    return df

def mark_empty_cells_excel(file_path, df, output_path):
    """在 Excel 文件中标记空单元格"""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 设置填充颜色为黄色
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row_index, row in df.iterrows():
        for col_index, value in row.items():
            if pd.isna(value):
                # Excel 的索引从 1 开始
                sheet.cell(row=row_index + 2, column=df.columns.get_loc(col_index) + 1).fill = fill

    workbook.save(output_path)

def clean_data(df, file_path):
    """清洗数据"""
    # 查找空数据并进行标注
    empty_cells = df.isnull()  # 返回一个布尔值的 DataFrame，空值为 True

    # 1. 标注空数据
    if file_path.endswith(('.xlsx', '.xls')):
        output_path = file_path.replace('.xlsx', '_marked.xlsx').replace('.xls', '_marked.xls')
        mark_empty_cells_excel(file_path, df, output_path)
        messagebox.showinfo("提示", f"已标注空数据的 Excel 文件保存在: {output_path}")
    elif file_path.endswith('.csv'):
        output_path = file_path.replace('.csv', '_marked.csv')
        df[empty_cells].to_csv(output_path, index=False) # 将标记后的数据框写入新的 CSV 文件
        messagebox.showinfo("提示", f"已标注空数据的 CSV 文件保存在: {output_path}")
    else:
        messagebox.showerror("错误", "不支持的文件类型")
        return

    # 2. 删除包含空数据的行
    cleaned_df = df.dropna()

    # 3. 输出清洗后的数据
    if file_path.endswith(('.xlsx', '.xls')):
        output_path = file_path.replace('.xlsx', '_cleaned.xlsx').replace('.xls', '_cleaned.xls')
        cleaned_df.to_excel(output_path, index=False)
    elif file_path.endswith('.csv'):
        output_path = file_path.replace('.csv', '_cleaned.csv')
        cleaned_df.to_csv(output_path, index=False)

    messagebox.showinfo("提示", f"清洗后的数据已保存至: {output_path}")

def main():
    """主函数"""
    file_path = choose_file()
    if not file_path:
        return

    df = load_data(file_path)
    if df is None:
        return

    # 检查数据是否与 Salesforce 对象的字段匹配 (这里只是一个示例，你需要根据实际情况修改)
    salesforce_fields = {
        "Account": ["Name", "AccountNumber", "Industry", "Phone", "Website"],
        "Contact": ["FirstName", "LastName", "Email", "Phone", "AccountId"],
        "Opportunity": ["Name", "StageName", "CloseDate", "Amount", "AccountId"]
        # ... 其他 Salesforce 对象和字段 ...
    }
    object_name = ""
    if set(df.columns).issuperset(salesforce_fields["Account"]):
        object_name = "Account"
    elif set(df.columns).issuperset(salesforce_fields["Contact"]):
        object_name = "Contact"
    elif set(df.columns).issuperset(salesforce_fields["Opportunity"]):
        object_name = "Opportunity"
    else:
        messagebox.showinfo("提示", "未找到匹配的 Salesforce 对象，将进行通用数据清洗。")

    if object_name != "":
        messagebox.showinfo("提示", f"检测到数据可能属于 Salesforce 的 {object_name} 对象。")

    clean_data(df, file_path)

if __name__ == "__main__":
    main()